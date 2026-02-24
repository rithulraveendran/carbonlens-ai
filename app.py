

# Placeholder data in this project must be replaced before deployment.
# Add your Firebase service account JSON to static/firebase_key.json.


import io
import csv
import json
import os
import re
import secrets
import smtplib
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps
from typing import Any, Dict, List, Optional

import firebase_admin
from PyPDF2 import PdfReader
from firebase_admin import credentials, firestore
from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

try:
    from groq import Groq
except ImportError:
    Groq = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SECRET_KEY = os.environ.get("SECRET_KEY", "CHANGE_ME_SECRET_KEY")
FIREBASE_KEY_PATH = os.environ.get(
    "FIREBASE_KEY_PATH",
    os.path.join(BASE_DIR, "static", "firebase_key.json"),
)
EMISSION_FACTORS_PATH = os.environ.get(
    "EMISSION_FACTORS_PATH",
    os.path.join(BASE_DIR, "static", "emission_factors.csv"),
)
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "").strip()
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "change-me@example.com").strip()
APP_PASSWORD = os.environ.get("APP_PASSWORD", "").replace(" ", "")
OTP_TTL_MINUTES = int(os.environ.get("OTP_TTL_MINUTES", "10"))
HIGH_EMISSION_ALERT_TCO2E = float(os.environ.get("HIGH_EMISSION_ALERT_TCO2E", "350"))
ALERT_EMAIL_COOLDOWN_MINUTES = int(os.environ.get("ALERT_EMAIL_COOLDOWN_MINUTES", "60"))

if not os.path.exists(FIREBASE_KEY_PATH):
    raise RuntimeError(f"FIREBASE_KEY_PATH does not exist: {FIREBASE_KEY_PATH}")


def _initialize_firebase() -> Any:
    with open(FIREBASE_KEY_PATH, "r", encoding="utf-8") as key_file:
        service_info = json.load(key_file)

    project_id = service_info.get("project_id")
    if not project_id:
        raise RuntimeError("Invalid Firebase service account file: missing project_id")

    if not firebase_admin._apps:
        cred = credentials.Certificate(FIREBASE_KEY_PATH)
        firebase_admin.initialize_app(cred)

    return firestore.client()


def _get_groq_client() -> Optional[Any]:
    if not GROQ_API_KEY or not Groq:
        return None
    return Groq(api_key=GROQ_API_KEY)


app = Flask(__name__)
app.config["SECRET_KEY"] = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_UPLOAD_MB", "20")) * 1024 * 1024

ALLOWED_EXTENSIONS = {"pdf", "xlsx", "xlsm", "csv"}
HIGH_RISK_THRESHOLD = 60
STAFF_ACCESS_CODE = os.environ.get("STAFF_ACCESS_CODE", "CHANGE_ME_STAFF_CODE").strip()
SUPPLIER_ACCESS_CODE = os.environ.get("SUPPLIER_ACCESS_CODE", "CHANGE_ME_SUPPLIER_CODE").strip()

db = _initialize_firebase()
groq_client = _get_groq_client()
client = groq_client


def _derive_access_role(user_data: Dict[str, Any]) -> str:
    role_raw = str((user_data or {}).get("access_role") or "").strip().lower()
    if role_raw in {"staff", "supplier"}:
        return role_raw
    legacy_role = str((user_data or {}).get("role") or "").strip().lower()
    return "supplier" if legacy_role == "supplier" else "staff"


def _resolve_signup_access(access_code: str) -> Optional[Dict[str, str]]:
    normalized = str(access_code or "").strip().upper()
    if normalized == STAFF_ACCESS_CODE:
        return {"access_role": "staff"}
    if normalized == SUPPLIER_ACCESS_CODE:
        return {"access_role": "supplier"}
    return None


def _generate_supplier_id() -> str:
    letters = "ABCDEFGHJKLMNPQRSTUVWXYZ"
    digits = "0123456789"
    attempts = 40
    for _ in range(attempts):
        candidate = f"{secrets.choice(letters)}{secrets.choice(letters)}{secrets.choice(digits)}{secrets.choice(digits)}"

        existing = db.collection("users").where("supplier_id", "==", candidate).limit(1).stream()
        if next(existing, None):
            continue
        existing_supplier_code = db.collection("suppliers").where("supplier_code", "==", candidate).limit(1).stream()
        if next(existing_supplier_code, None):
            continue
        return candidate

    raise RuntimeError("Unable to generate a unique supplier ID. Please try again.")


def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if "user_email" not in session:
            flash("Please log in to continue.", "error")
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapped


def parse_timestamp(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def normalize_activity_key(activity_key: str) -> str:
    key = str(activity_key or "").strip().lower()
    key = key.replace("-", "_").replace(" ", "_")
    key = re.sub(r"[^a-z0-9_]", "", key)
    key = re.sub(r"_+", "_", key).strip("_")
    return key


def load_emission_factors() -> Dict[str, Dict[str, Any]]:
    if not os.path.exists(EMISSION_FACTORS_PATH):
        raise RuntimeError(f"Emission factors file not found: {EMISSION_FACTORS_PATH}")

    factors: Dict[str, Dict[str, Any]] = {}
    with open(EMISSION_FACTORS_PATH, "r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        required_columns = {"activity_type", "unit", "emission_factor_kg_co2e", "region", "source"}
        if not reader.fieldnames or not required_columns.issubset({name.strip() for name in reader.fieldnames}):
            raise RuntimeError("Invalid emission_factors.csv format.")

        for row in reader:
            activity_type = normalize_activity_key(row.get("activity_type", ""))
            factor_raw = str(row.get("emission_factor_kg_co2e", "")).strip()
            if not activity_type or not factor_raw:
                continue

            try:
                factor_value = float(factor_raw)
            except Exception:
                continue

            factors[activity_type] = {
                "factor": factor_value,
                "unit": str(row.get("unit", "")).strip(),
                "region": str(row.get("region", "")).strip(),
                "source": str(row.get("source", "")).strip(),
            }

    if not factors:
        raise RuntimeError("No valid emission factors loaded from emission_factors.csv.")

    return factors


def _coerce_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return default


def normalize_reduction_suggestions(value: Any, max_items: int = 8) -> List[str]:
    if isinstance(value, str):
        candidates = re.split(r"[\r\n]+", value)
    elif isinstance(value, (list, tuple, set)):
        candidates = [str(item) for item in value]
    else:
        candidates = []

    cleaned: List[str] = []
    for item in candidates:
        line = str(item or "").strip()
        line = re.sub(r"^[\-\*\u2022\d\.\)\s]+", "", line)
        if not line:
            continue
        if line not in cleaned:
            cleaned.append(line)
        if len(cleaned) >= max_items:
            break
    return cleaned


def resolve_activity_type(activity_key: str) -> Optional[str]:
    normalized = normalize_activity_key(activity_key)
    if normalized in EMISSION_FACTORS:
        return normalized

    aliases = {
        "diesel": "diesel_litre",
        "diesel_litres": "diesel_litre",
        "diesel_liters": "diesel_litre",
        "petrol": "petrol_litre",
        "petrol_litres": "petrol_litre",
        "petrol_liters": "petrol_litre",
        "electricity": "electricity_kwh",
        "electricity_units": "electricity_kwh",
        "power_kwh": "electricity_kwh",
        "steel_tonne": "steel_ton",
        "steel_tonnes": "steel_ton",
        "steel_tons": "steel_ton",
        "freight_tonkm": "freight_ton_km",
        "freight_tkm": "freight_ton_km",
        "ton_km_freight": "freight_ton_km",
        "rail_freight_tonkm": "rail_freight_ton_km",
        "rail_ton_km": "rail_freight_ton_km",
        "raw_material_kg": "raw_materials_kg",
        "raw_materials": "raw_materials_kg",
    }

    resolved = aliases.get(normalized)
    if resolved and resolved in EMISSION_FACTORS:
        return resolved

    singular = normalized[:-1] if normalized.endswith("s") else normalized
    if singular in EMISSION_FACTORS:
        return singular

    return None


def calculate_emissions(activity_data: Dict[str, Any]) -> Dict[str, Any]:
    if not activity_data or not isinstance(activity_data, dict):
        raise ValueError("No activity data found for deterministic emissions calculation.")

    total_kg = 0.0
    missing_activity_types: List[str] = []
    normalized_activity_data: Dict[str, float] = {}
    breakdown: List[Dict[str, Any]] = []

    for activity_key, quantity in activity_data.items():
        resolved_key = resolve_activity_type(str(activity_key))
        if not resolved_key:
            missing_activity_types.append(str(activity_key))
            continue

        factor_info = EMISSION_FACTORS.get(resolved_key)
        if not factor_info:
            missing_activity_types.append(str(activity_key))
            continue

        value = _coerce_float(quantity, 0.0)
        if value <= 0:
            continue

        factor = _coerce_float(factor_info.get("factor"), 0.0)
        emission_kg = value * factor
        total_kg += emission_kg
        normalized_activity_data[resolved_key] = normalized_activity_data.get(resolved_key, 0.0) + value

        breakdown.append(
            {
                "activity_type": resolved_key,
                "input_value": value,
                "input_unit": factor_info.get("unit", ""),
                "emission_factor_kg_co2e": factor,
                "emissions_kg_co2e": round(emission_kg, 6),
                "region": factor_info.get("region", ""),
                "source": factor_info.get("source", ""),
            }
        )

    if missing_activity_types:
        missing_csv = ", ".join(sorted(set(missing_activity_types)))
        raise ValueError(f"Activity type(s) not found in emission_factors.csv: {missing_csv}")

    if not breakdown:
        raise ValueError("No valid positive activity values found for deterministic emissions calculation.")

    total_tco2e = total_kg / 1000.0
    return {
        "total_kg_co2e": round(total_kg, 6),
        "total_tco2e": round(total_tco2e, 6),
        "activity_data": normalized_activity_data,
        "breakdown": breakdown,
    }


EMISSION_FACTORS = load_emission_factors()


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def risk_level(score: float) -> str:
    if score >= 80:
        return "Very High Risk"
    if score >= 60:
        return "High Risk"
    if score >= 40:
        return "Moderate Risk"
    return "Low Risk"


def emissions_risk_band(total_emissions_tco2e: float) -> str:
    if total_emissions_tco2e < 50:
        return "Negligible Risk"
    if total_emissions_tco2e < 150:
        return "Low Risk"
    if total_emissions_tco2e < 350:
        return "Moderate Risk"
    if total_emissions_tco2e < 700:
        return "High Risk"
    return "Very High Risk"


def otp_purpose_title(context: str) -> str:
    titles = {
        "signup": "Account Signup",
        "signin": "Sign In",
        "reset": "Password Reset",
    }
    return titles.get(context, "Verification")


def generate_otp_code() -> str:
    return f"{secrets.randbelow(1_000_000):06d}"


def set_otp_session(context: str, email: str) -> str:
    code = generate_otp_code()
    session["otp_context"] = context
    session["otp_email"] = email
    session["otp_code"] = code
    session["otp_expires_at"] = (datetime.now(timezone.utc) + timedelta(minutes=OTP_TTL_MINUTES)).isoformat()
    return code


def clear_otp_session() -> None:
    for key in ["otp_context", "otp_email", "otp_code", "otp_expires_at"]:
        session.pop(key, None)


def otp_expired() -> bool:
    expires_raw = session.get("otp_expires_at")
    if not expires_raw:
        return True
    try:
        return datetime.now(timezone.utc) > datetime.fromisoformat(expires_raw)
    except Exception:
        return True


def send_otp_email(recipient_email: str, otp_code: str, purpose: str) -> bool:
    if not SENDER_EMAIL or not APP_PASSWORD:
        return False

    subject = f"CarbonLens AI OTP - {otp_purpose_title(purpose)}"
    body_text = (
        f"Your CarbonLens AI OTP for {otp_purpose_title(purpose)} is: {otp_code}\n\n"
        f"This code expires in {OTP_TTL_MINUTES} minutes.\n"
        "If you did not request this, please ignore this email."
    )
    body_html = f"""
    <html>
      <body style="margin:0;padding:0;background:#f6f8fb;font-family:Arial,sans-serif;color:#0f172a;">
        <table width="100%" cellpadding="0" cellspacing="0" style="padding:24px 12px;">
          <tr>
            <td align="center">
              <table width="560" cellpadding="0" cellspacing="0" style="background:#ffffff;border:1px solid #e2e8f0;border-radius:14px;overflow:hidden;">
                <tr>
                  <td style="background:#0f172a;padding:18px 24px;color:#ffffff;font-size:18px;font-weight:700;">
                    CarbonLens AI
                  </td>
                </tr>
                <tr>
                  <td style="padding:24px;">
                    <p style="margin:0 0 10px 0;font-size:14px;color:#475569;">{otp_purpose_title(purpose)}</p>
                    <p style="margin:0 0 16px 0;font-size:16px;font-weight:700;">Your verification code</p>
                    <div style="display:inline-block;padding:10px 16px;border:1px dashed #7F00FF;border-radius:10px;background:#f5f3ff;color:#6d28d9;font-size:28px;font-weight:800;letter-spacing:4px;">
                      {otp_code}
                    </div>
                    <p style="margin:18px 0 0 0;font-size:13px;color:#64748b;">This code expires in {OTP_TTL_MINUTES} minutes.</p>
                    <p style="margin:10px 0 0 0;font-size:12px;color:#94a3b8;">If you did not request this, you can safely ignore this email.</p>
                  </td>
                </tr>
              </table>
            </td>
          </tr>
        </table>
      </body>
    </html>
    """.strip()

    message = MIMEMultipart("alternative")
    message["Subject"] = subject
    message["From"] = SENDER_EMAIL
    message["To"] = recipient_email
    message.attach(MIMEText(body_text, "plain"))
    message.attach(MIMEText(body_html, "html"))

    try:
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=20) as smtp:
            smtp.starttls()
            smtp.login(SENDER_EMAIL, APP_PASSWORD)
            smtp.sendmail(SENDER_EMAIL, [recipient_email], message.as_string())
        return True
    except Exception:
        return False


def send_high_emission_alert_email(
    recipient_email: str,
    supplier_name: str,
    total_emissions_tco2e: float,
    threshold_tco2e: float,
) -> bool:
    if not SENDER_EMAIL or not APP_PASSWORD:
        return False

    subject = "CarbonLens AI Alert - High Supplier Emissions"
    body_text = (
        f"Supplier '{supplier_name}' is above the high-emission threshold.\n\n"
        f"Current total emissions: {total_emissions_tco2e:,.2f} tCO2e\n"
        f"Configured threshold: {threshold_tco2e:,.2f} tCO2e\n\n"
        "Review this supplier in CarbonLens AI for mitigation actions."
    )
    body_html = f"""
    <html>
      <body style="margin:0;padding:0;background:#f6f8fb;font-family:Arial,sans-serif;color:#0f172a;">
        <table width="100%" cellpadding="0" cellspacing="0" style="padding:24px 12px;">
          <tr>
            <td align="center">
              <table width="560" cellpadding="0" cellspacing="0" style="background:#ffffff;border:1px solid #e2e8f0;border-radius:14px;overflow:hidden;">
                <tr>
                  <td style="background:#7f1d1d;padding:18px 24px;color:#ffffff;font-size:18px;font-weight:700;">
                    High Emissions Alert
                  </td>
                </tr>
                <tr>
                  <td style="padding:24px;">
                    <p style="margin:0 0 10px 0;font-size:14px;color:#475569;">Supplier monitoring notification</p>
                    <p style="margin:0 0 16px 0;font-size:16px;font-weight:700;">{supplier_name}</p>
                    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;">
                      <tr>
                        <td style="padding:10px;border:1px solid #e2e8f0;font-size:13px;color:#64748b;">Current total</td>
                        <td style="padding:10px;border:1px solid #e2e8f0;font-size:13px;font-weight:700;color:#0f172a;">{total_emissions_tco2e:,.2f} tCO2e</td>
                      </tr>
                      <tr>
                        <td style="padding:10px;border:1px solid #e2e8f0;font-size:13px;color:#64748b;">Threshold</td>
                        <td style="padding:10px;border:1px solid #e2e8f0;font-size:13px;font-weight:700;color:#0f172a;">{threshold_tco2e:,.2f} tCO2e</td>
                      </tr>
                    </table>
                    <p style="margin:16px 0 0 0;font-size:13px;color:#64748b;">Recommended: review supplier emissions trend and assign mitigation actions.</p>
                  </td>
                </tr>
              </table>
            </td>
          </tr>
        </table>
      </body>
    </html>
    """.strip()

    message = MIMEMultipart("alternative")
    message["Subject"] = subject
    message["From"] = SENDER_EMAIL
    message["To"] = recipient_email
    message.attach(MIMEText(body_text, "plain"))
    message.attach(MIMEText(body_html, "html"))

    try:
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=20) as smtp:
            smtp.starttls()
            smtp.login(SENDER_EMAIL, APP_PASSWORD)
            smtp.sendmail(SENDER_EMAIL, [recipient_email], message.as_string())
        return True
    except Exception:
        app.logger.exception("Failed to send high-emission alert email to %s", recipient_email)
        return False


def _collect_alert_recipients(supplier: Dict[str, Any], default_recipient_email: str) -> List[str]:
    recipients: set[str] = set()
    default_email = str(default_recipient_email or "").strip().lower()
    if default_email:
        recipients.add(default_email)

    supplier_owner_email = str(supplier.get("user_email", "") or "").strip().lower()
    if supplier_owner_email:
        recipients.add(supplier_owner_email)

    supplier_code_raw = str(supplier.get("supplier_code", "") or "").strip()
    if supplier_code_raw:
        lookup_codes = {supplier_code_raw}
        lookup_codes.add(supplier_code_raw.upper())
        lookup_codes.add(supplier_code_raw.lower())

        for code in lookup_codes:
            for user_doc in db.collection("users").where("supplier_id", "==", code).stream():
                data = user_doc.to_dict() or {}
                email = str(data.get("email", user_doc.id) or "").strip().lower()
                if email:
                    recipients.add(email)

            profile_doc = db.collection("supplier_profiles").document(code).get()
            if profile_doc.exists:
                pdata = profile_doc.to_dict() or {}
                email = str(pdata.get("user_email", "") or "").strip().lower()
                if email:
                    recipients.add(email)

    return sorted(recipients)


def trigger_high_emission_alert_if_needed(
    supplier: Dict[str, Any],
    recipient_email: str,
    incremental_emissions: float,
) -> None:
    supplier_old_total = _coerce_float(supplier.get("total_emissions"), 0.0)
    supplier_new_total = supplier_old_total + _coerce_float(incremental_emissions, 0.0)
    should_alert = supplier_new_total >= HIGH_EMISSION_ALERT_TCO2E
    if not should_alert:
        return

    cooldown_ok = True
    last_alert_raw = str(supplier.get("last_emission_alert_at", "") or "").strip()
    if last_alert_raw:
        try:
            last_alert_at = datetime.fromisoformat(last_alert_raw)
            if last_alert_at.tzinfo is None:
                last_alert_at = last_alert_at.replace(tzinfo=timezone.utc)
            elapsed_minutes = (datetime.now(timezone.utc) - last_alert_at).total_seconds() / 60.0
            cooldown_ok = elapsed_minutes >= ALERT_EMAIL_COOLDOWN_MINUTES
        except Exception:
            cooldown_ok = True

    if not cooldown_ok:
        return

    recipients = _collect_alert_recipients(supplier, recipient_email)
    if not recipients:
        return

    sent_any = False
    for email in recipients:
        sent = send_high_emission_alert_email(
            recipient_email=email,
            supplier_name=supplier.get("supplier_name", supplier.get("name", "Supplier")),
            total_emissions_tco2e=supplier_new_total,
            threshold_tco2e=HIGH_EMISSION_ALERT_TCO2E,
        )
        sent_any = sent_any or sent

    if sent_any:
        db.collection("suppliers").document(str(supplier.get("id", ""))).update(
            {
                "last_emission_alert_at": datetime.now(timezone.utc).isoformat(),
                "last_emission_alert_total": round(supplier_new_total, 6),
            }
        )
    else:
        app.logger.warning(
            "High-emission alert email not sent for supplier %s (recipients: %s).",
            supplier.get("id", ""),
            ", ".join(recipients),
        )


def get_user_profile(user_email: str) -> Optional[Dict[str, Any]]:
    user_doc = db.collection("users").document(user_email).get()
    if not user_doc.exists:
        return None
    data = user_doc.to_dict() or {}
    access_role = _derive_access_role(data)
    return {
        "email": user_email,
        "name": data.get("name", "User"),
        "company_name": data.get("company_name", ""),
        "role": data.get("role", "Staff" if access_role == "staff" else "Supplier"),
        "access_role": access_role,
        "supplier_id": str(data.get("supplier_id", "") or "").strip(),
        "password": data.get("password", ""),
    }


def _shape_supplier_data(data: Dict[str, Any], supplier_id: str) -> Dict[str, Any]:
    data = dict(data or {})
    data["id"] = supplier_id
    data["total_emissions"] = float(data.get("total_emissions", 0) or 0)
    data["coverage_percent"] = float(data.get("coverage_percent", 0) or 0)
    data["risk_score"] = calculate_risk_score(data["total_emissions"], data["coverage_percent"])
    data["emissions_tco2e"] = data["total_emissions"]
    data["coverage_pct"] = round(data["coverage_percent"], 2)
    data["category"] = data.get("sector", "General")
    data["name"] = data.get("supplier_name", "Supplier")
    data["region"] = data.get("region", "Global")
    data["normalization_alert"] = "Units standardized to tCO2e."
    data["ai_insight"] = data.get(
        "ai_insight",
        f"Supplier-ID {supplier_id}: Scope 3 profile built from processed documents.",
    )
    return data


def get_supplier_by_access_id(supplier_access_id: str) -> Optional[Dict[str, Any]]:
    supplier_access_id = str(supplier_access_id or "").strip()
    if not supplier_access_id:
        return None

    code_docs = db.collection("suppliers").where("supplier_code", "==", supplier_access_id).limit(1).stream()
    for doc in code_docs:
        return _shape_supplier_data(doc.to_dict() or {}, doc.id)

    direct_doc = db.collection("suppliers").document(supplier_access_id).get()
    if direct_doc.exists:
        return _shape_supplier_data(direct_doc.to_dict() or {}, direct_doc.id)
    return None


def _resolve_company_name(owner_email: str) -> str:
    owner_email = str(owner_email or "").strip()
    if not owner_email:
        return "Unknown Company"
    owner_doc = db.collection("users").document(owner_email).get()
    if owner_doc.exists:
        owner_data = owner_doc.to_dict() or {}
        resolved = str(owner_data.get("company_name", "") or "").strip()
        if resolved:
            return resolved
    return owner_email


def list_supplier_company_records(supplier_external_id: str) -> List[Dict[str, Any]]:
    supplier_external_id = str(supplier_external_id or "").strip()
    if not supplier_external_id:
        return []

    records: Dict[str, Dict[str, Any]] = {}

    for doc in db.collection("suppliers").where("supplier_code", "==", supplier_external_id).stream():
        records[doc.id] = _shape_supplier_data(doc.to_dict() or {}, doc.id)

    legacy = get_supplier_by_access_id(supplier_external_id)
    if legacy:
        legacy_id = str(legacy.get("id", "")).strip()
        if legacy_id and legacy_id not in records:
            records[legacy_id] = legacy

    return sorted(records.values(), key=lambda item: item.get("total_emissions", 0.0), reverse=True)


def list_supplier_company_options(supplier_external_id: str) -> List[Dict[str, Any]]:
    options: List[Dict[str, Any]] = []
    for record in list_supplier_company_records(supplier_external_id):
        owner_email = str(record.get("user_email", "") or "").strip()
        company_name = _resolve_company_name(owner_email)
        options.append(
            {
                "supplier_id": str(record.get("id", "")).strip(),
                "owner_email": owner_email,
                "company_name": company_name,
                "supplier_name": str(record.get("name", "Supplier")),
                "supplier_code": str(record.get("supplier_code", "") or "").strip(),
                "supplier": record,
            }
        )
    return options


def list_documents_for_supplier(supplier_id: str, owner_email: Optional[str] = None) -> List[Dict[str, Any]]:
    documents: List[Dict[str, Any]] = []
    query = db.collection("documents").where("supplier_id", "==", supplier_id)
    if owner_email:
        query = query.where("user_email", "==", owner_email)

    for doc in query.stream():
        data = doc.to_dict() or {}
        data["id"] = doc.id
        data["created_at_iso"] = parse_timestamp(data.get("created_at"))
        data["extracted_emissions"] = float(data.get("extracted_emissions", 0) or 0)
        data["confidence_percent"] = float(data.get("confidence_percent", 0) or 0)
        documents.append(data)

    documents.sort(key=lambda item: item.get("created_at_iso", ""), reverse=True)
    return documents


def list_user_suppliers(user_email: str) -> List[Dict[str, Any]]:
    suppliers = []
    docs = db.collection("suppliers").where("user_email", "==", user_email).stream()
    for doc in docs:
        suppliers.append(_shape_supplier_data(doc.to_dict() or {}, doc.id))

    suppliers.sort(key=lambda item: item.get("risk_score", 0), reverse=True)
    return suppliers


def get_supplier_owned(supplier_id: str, user_email: str) -> Optional[Dict[str, Any]]:
    supplier_ref = db.collection("suppliers").document(supplier_id)
    supplier_doc = supplier_ref.get()
    if not supplier_doc.exists:
        return None

    data = supplier_doc.to_dict() or {}
    if data.get("user_email") != user_email:
        return None

    data = _shape_supplier_data(data, supplier_doc.id)
    data["last_updated"] = parse_timestamp(data.get("created_at"))[:10]
    return data


def list_documents(user_email: str, supplier_id: Optional[str] = None) -> List[Dict[str, Any]]:
    documents: List[Dict[str, Any]] = []
    query = db.collection("documents").where("user_email", "==", user_email)
    if supplier_id:
        query = query.where("supplier_id", "==", supplier_id)

    for doc in query.stream():
        data = doc.to_dict() or {}
        data["id"] = doc.id
        data["created_at_iso"] = parse_timestamp(data.get("created_at"))
        data["extracted_emissions"] = float(data.get("extracted_emissions", 0) or 0)
        data["confidence_percent"] = float(data.get("confidence_percent", 0) or 0)
        data["scope_type"] = str(data.get("scope_type", "scope3") or "scope3").strip().lower()
        documents.append(data)

    documents.sort(key=lambda item: item.get("created_at_iso", ""), reverse=True)

    return documents


def _verification_label(status: str) -> str:
    status_key = str(status or "").strip().lower()
    status_map = {
        "under_ai_processing": "Under AI Processing",
        "waiting_for_staff_approval": "Waiting for Staff Approval",
        "verified": "Verified",
        "rejected": "Rejected",
    }
    return status_map.get(status_key, "Verified")


def _build_verification_queue(user_email: str) -> List[Dict[str, Any]]:
    pending_statuses = {"under_ai_processing", "waiting_for_staff_approval"}
    queue: List[Dict[str, Any]] = []
    for item in list_documents(user_email):
        if str(item.get("scope_type", "scope3")).strip().lower() != "scope3":
            continue
        status = str(item.get("verification_status", "verified")).strip().lower()
        if status not in pending_statuses:
            continue
        activity_data = item.get("activity_data") if isinstance(item.get("activity_data"), dict) else {}
        queue.append(
            {
                "id": item.get("id", ""),
                "supplier_id": item.get("supplier_id", ""),
                "supplier_name": item.get("supplier_name", "Supplier"),
                "filename": item.get("original_filename", "Document"),
                "status_label": _verification_label(status),
                "created_at": item.get("created_at_iso", "")[:10],
                "ai_summary": item.get("ai_summary", "AI summary unavailable."),
                "ai_flag_reason": item.get(
                    "ai_flag_reason",
                    "AI flagged this upload for staff verification before final score publication.",
                ),
                "activity_data": activity_data,
            }
        )
    return queue


def _build_supplier_data_summary(user_email: str, supplier_id: str) -> str:
    user = get_user_profile(user_email) or {}
    company_options = list_supplier_company_options(supplier_id)
    if not company_options:
        return (
            f"User: {user.get('name', 'User')} | Role: Supplier\n"
            f"My Supplier ID: {supplier_id or 'N/A'}\n"
            "No linked company records found for this supplier ID."
        )

    total_emissions = sum(_coerce_float(item["supplier"].get("total_emissions"), 0.0) for item in company_options)
    total_docs = 0
    lines = [
        f"User: {user.get('name', 'User')} | Role: Supplier",
        f"My Supplier ID: {supplier_id}",
        f"Linked companies: {len(company_options)}",
        f"My combined Scope 3 total: {total_emissions:,.2f} tCO2e",
    ]
    for option in company_options:
        supplier = option["supplier"]
        documents = list_documents_for_supplier(option["supplier_id"], owner_email=option["owner_email"])
        total_docs += len(documents)
        lines.append(
            f"Company: {option['company_name']} | Supplier Record: {supplier.get('name', 'Supplier')} "
            f"(ID: {option['supplier_id']}) | "
            f"{_coerce_float(supplier.get('total_emissions'), 0.0):,.2f} tCO2e | "
            f"Coverage {_coerce_float(supplier.get('coverage_percent'), 0.0):.2f}% | Docs {len(documents)}"
        )
        for item in documents[:4]:
            lines.append(
                f"{item.get('original_filename', 'N/A')} | {item.get('created_at_iso', '')[:10]} | "
                f"{_verification_label(str(item.get('verification_status', 'verified')))} | "
                f"{_coerce_float(item.get('extracted_emissions'), 0.0):,.4f} tCO2e"
            )
    lines.insert(3, f"Recent submissions: {total_docs}")
    return "\n".join(lines)


def build_user_data_summary(user_email: str, access_role: str = "staff", supplier_id: str = "") -> str:
    if str(access_role).lower() == "supplier":
        return _build_supplier_data_summary(user_email, supplier_id)

    user = get_user_profile(user_email) or {}
    suppliers = list_user_suppliers(user_email)
    documents = list_documents(user_email)

    if not suppliers:
        return (
            f"User: {user.get('name', 'User')} | Company: {user.get('company_name', 'N/A')}\n"
            "User has 0 suppliers and no supplier records to analyze."
        )

    total_emissions = sum(_coerce_float(item.get("total_emissions"), 0.0) for item in suppliers)
    avg_coverage = (
        sum(_coerce_float(item.get("coverage_percent"), 0.0) for item in suppliers) / len(suppliers)
        if suppliers
        else 0.0
    )
    lines: List[str] = [
        f"User: {user.get('name', 'User')} | Company: {user.get('company_name', 'N/A')}",
        f"Portfolio Summary: Suppliers {len(suppliers)} | Documents {len(documents)} | Total Scope 3 {total_emissions:,.2f} tCO2e | Avg Coverage {avg_coverage:.2f}%",
        f"User has {len(suppliers)} suppliers:",
    ]
    for supplier in suppliers:
        lines.append(
            f"{supplier.get('name', 'Supplier')} (ID: {supplier.get('id', '')}) - "
            f"{supplier.get('category', 'General')} - {supplier.get('region', 'Global')} - "
            f"{_coerce_float(supplier.get('total_emissions'), 0.0):,.2f} tCO2e - "
            f"{risk_level(_coerce_float(supplier.get('risk_score'), 0.0))} - "
            f"Coverage {_coerce_float(supplier.get('coverage_percent'), 0.0):.2f}%"
        )

    if documents:
        lines.append("Recent documents:")
        for item in documents[:10]:
            lines.append(
                f"{item.get('original_filename', 'N/A')} | Supplier {item.get('supplier_name', 'N/A')} | "
                f"{_coerce_float(item.get('extracted_emissions'), 0.0):,.4f} tCO2e | "
                f"{item.get('category', 'General')} | Confidence {_coerce_float(item.get('confidence_percent'), 0.0):.0f}%"
            )

    return "\n".join(lines)


def get_carbonlens_ai_response(
    user_message: str,
    chat_history: List[Dict[str, Any]],
    user_email: str,
    access_role: str = "staff",
    supplier_id: str = "",
) -> str:
    data_context = build_user_data_summary(user_email, access_role=access_role, supplier_id=supplier_id)
    if not data_context.strip():
        return "I don't have enough data to answer that."

    role_prompt = (
        "Help the auditor identify high-risk suppliers and audit discrepancies."
        if str(access_role).lower() == "staff"
        else "Act as a sustainability consultant. Suggest energy-saving measures based on the user's recent uploads."
    )

    system_prompt = f"""
You are CarbonLens AI, an ESG intelligence assistant for enterprise users.
Role mode: {access_role}
Role directive: {role_prompt}

Use only the provided CarbonLens database context to answer.
Primary scope:
- Suppliers
- Scope 3 emissions
- Emission breakdown
- Risk scores
- Coverage percentages
- Uploaded documents
- ESG interpretation
- Topic-related sustainability guidance (decarbonization, reporting, supplier engagement, target setting)

Rules:
1. Never fabricate numbers, suppliers, or documents.
2. If a database-specific answer is requested and data is missing, respond exactly: I don't have enough data to answer that.
3. Respond naturally to greetings and simple conversational messages (e.g., hi, hello, thanks) in a short friendly way.
4. For topic-related questions that do not require exact database values, provide practical ESG guidance without inventing internal data.
5. For clearly unrelated or disallowed topics (politics, coding help, medical advice, general trivia), politely redirect to CarbonLens ESG scope.
6. Keep responses short, professional, and structured.
7. When comparing suppliers, use bullet points.
8. Follow the role directive strictly when advising.

Database context:
{data_context}
""".strip()

    messages_for_model: List[Dict[str, str]] = [{"role": "system", "content": system_prompt}]

    history_items = chat_history if isinstance(chat_history, list) else []
    for msg in history_items:
        if not isinstance(msg, dict):
            continue
        role = str(msg.get("role", "")).strip()
        content = str(msg.get("content", "")).strip()
        if role in {"user", "assistant"} and content:
            messages_for_model.append({"role": role, "content": content})

    messages_for_model.append({"role": "user", "content": str(user_message).strip()})

    if not client:
        return "CarbonLens AI chat is currently unavailable because the Groq client is not configured."

    try:
        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=messages_for_model,
        )
        return response.choices[0].message.content.strip()
    except Exception as exc:
        return f"Error connecting to CarbonLens AI: {str(exc)}"

def extract_text_from_pdf(file_bytes: bytes) -> str:
    if not file_bytes:
        return ""

    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        extracted = []
        for page in reader.pages:
            extracted.append(page.extract_text() or "")
        return "\n".join(extracted).strip()
    except Exception:
        return ""


def extract_activity_data_from_table_file(file_bytes: bytes, filename: str) -> Dict[str, float]:
    if not file_bytes:
        return {}

    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    rows: List[List[Any]] = []

    if extension == "csv":
        decoded = file_bytes.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(decoded))
        rows = [list(row) for row in reader]
    elif extension in {"xlsx", "xlsm"}:
        if not load_workbook:
            raise RuntimeError("Excel upload requires openpyxl. Install it with: pip install openpyxl")
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    else:
        return {}

    rows = [row for row in rows if any(str(cell or "").strip() for cell in row)]
    if not rows:
        return {}

    activity_data: Dict[str, float] = {}
    headers = [normalize_activity_key(str(col or "")) for col in rows[0]]

    def add_activity_value(activity_key: Any, value: Any) -> None:
        resolved = resolve_activity_type(str(activity_key))
        numeric_value = _coerce_float(value, 0.0)
        if resolved and numeric_value > 0:
            activity_data[resolved] = activity_data.get(resolved, 0.0) + numeric_value

    header_to_index = {name: idx for idx, name in enumerate(headers) if name}
    activity_idx = header_to_index.get("activity_type")
    quantity_idx = (
        header_to_index.get("quantity")
        if "quantity" in header_to_index
        else header_to_index.get("value", header_to_index.get("amount"))
    )

    if activity_idx is not None and quantity_idx is not None:
        for row in rows[1:]:
            if activity_idx < len(row) and quantity_idx < len(row):
                add_activity_value(row[activity_idx], row[quantity_idx])

    for idx, header in enumerate(headers):
        resolved_header = resolve_activity_type(header)
        if not resolved_header:
            continue
        for row in rows[1:]:
            if idx < len(row):
                numeric_value = _coerce_float(row[idx], 0.0)
                if numeric_value > 0:
                    activity_data[resolved_header] = activity_data.get(resolved_header, 0.0) + numeric_value

    for row in rows[1:]:
        if len(row) >= 2:
            add_activity_value(row[0], row[1])

    return activity_data


def infer_category(text: str) -> str:
    lowered = text.lower()
    if any(word in lowered for word in ["freight", "shipping", "transport", "logistics"]):
        return "Logistics"
    if any(word in lowered for word in ["material", "procurement", "purchased goods", "supplier"]):
        return "Purchased Goods"
    if any(word in lowered for word in ["travel", "flight", "hotel"]):
        return "Business Travel"
    if any(word in lowered for word in ["waste", "recycling", "disposal"]):
        return "Waste"
    return "General"


def fallback_regex_extraction(document_text: str) -> Dict[str, Any]:
    text = document_text or ""

    supplier_match = re.search(r"(?:supplier|vendor|company)\s*[:\-]\s*([A-Za-z0-9 &,.-]{2,80})", text, re.IGNORECASE)
    supplier_name = supplier_match.group(1).strip() if supplier_match else "Unknown Supplier"
    
    def extract_quantity(patterns: List[str]) -> float:
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                quantity = _coerce_float(match.group(1), 0.0)
                if quantity > 0:
                    return quantity
        return 0.0

    activity_data: Dict[str, float] = {}
    pattern_map = {
        "diesel_litre": [
            r"(?:diesel(?:\s+consumption|\s+usage)?)[^\d]{0,25}([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:litre|litres|liter|liters)\b",
        ],
        "petrol_litre": [
            r"(?:petrol(?:\s+consumption|\s+usage)?|gasoline)[^\d]{0,25}([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:litre|litres|liter|liters)\b",
        ],
        "electricity_kwh": [
            r"(?:electricity|power)(?:\s+consumption|\s+usage)?[^\d]{0,25}([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:kwh|kw\s*h|units)\b",
        ],
        "steel_ton": [
            r"(?:steel(?:\s+used|\s+consumed)?)[^\d]{0,25}([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:ton|tons|tonne|tonnes|t)\b",
        ],
        "freight_ton_km": [
            r"(?:freight|road\s+transport|shipping)[^\d]{0,25}([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:ton[\s_-]?km|tonkm|tkm)\b",
        ],
        "rail_freight_ton_km": [
            r"(?:rail\s+freight|rail\s+transport)[^\d]{0,25}([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:ton[\s_-]?km|tonkm|tkm)\b",
        ],
        "raw_materials_kg": [
            r"(?:raw\s+materials?)[^\d]{0,25}([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:kg|kilogram|kilograms)\b",
        ],
    }

    for activity_type, patterns in pattern_map.items():
        quantity = extract_quantity(patterns)
        if quantity > 0:
            activity_data[activity_type] = quantity

    category = infer_category(text)
    confidence = 68 if activity_data else 45
    summary_text = (
        "Fallback extraction applied. "
        f"Detected {len(activity_data)} activity metric(s) for deterministic factor-based calculation."
    )

    return {
        "supplier_name": supplier_name,
        "activity_data": activity_data,
        "raw_activity_data": activity_data.copy(),
        "category": category,
        "confidence_percent": confidence,
        "summary_text": summary_text,
    }


def _extract_json_payload(content: str) -> Dict[str, Any]:
    try:
        return json.loads(content)
    except Exception:
        pass

    json_match = re.search(r"\{.*\}", content, re.DOTALL)
    if json_match:
        try:
            return json.loads(json_match.group(0))
        except Exception:
            return {}
    return {}


def extract_emissions_with_ai(document_text: str) -> Dict[str, Any]:
    if groq_client and document_text.strip():
        activity_types = ", ".join(sorted(EMISSION_FACTORS.keys()))
        prompt = (
            "Extract activity quantities from this supplier document for Scope 3 calculation. "
            f"Allowed activity_type keys: {activity_types}. "
            "Return strict JSON with keys: supplier_name, activity_data, category, confidence_percent, summary_text. "
            "activity_data must be an object where keys are activity types and values are numeric quantities."
        )

        try:
            response = groq_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {
                        "role": "system",
                        "content": "You are CarbonLensAI extraction engine. Output strict JSON only.",
                    },
                    {
                        "role": "user",
                        "content": f"{prompt}\n\nDocument content:\n{document_text[:12000]}",
                    },
                ],
                temperature=0,
            )
            content = (response.choices[0].message.content or "").strip()
            parsed = _extract_json_payload(content)
            if parsed:
                normalized_activity_data: Dict[str, float] = {}
                source_data = parsed.get("activity_data") if isinstance(parsed.get("activity_data"), dict) else parsed
                for key, value in (source_data or {}).items():
                    resolved = resolve_activity_type(str(key))
                    numeric_value = _coerce_float(value, 0.0)
                    if resolved and numeric_value > 0:
                        normalized_activity_data[resolved] = normalized_activity_data.get(resolved, 0.0) + numeric_value

                if not normalized_activity_data:
                    return fallback_regex_extraction(document_text)

                return {
                    "supplier_name": str(parsed.get("supplier_name") or "Unknown Supplier")[:120],
                    "activity_data": normalized_activity_data,
                    "raw_activity_data": normalized_activity_data.copy(),
                    "category": str(parsed.get("category") or infer_category(document_text)),
                    "confidence_percent": int(clamp(_coerce_float(parsed.get("confidence_percent"), 60), 1, 100)),
                    "summary_text": str(parsed.get("summary_text") or "AI extraction completed."),
                }
        except Exception:
            pass

    return fallback_regex_extraction(document_text)


def calculate_risk_score(total_emissions_tco2e: float, coverage_percent: float = 0.0) -> float:
    emissions = _coerce_float(total_emissions_tco2e, 0.0)
    coverage = clamp(_coerce_float(coverage_percent, 0.0), 0, 100)

    if emissions < 50:
        base_score = 20.0
    elif emissions < 150:
        base_score = 40.0
    elif emissions < 350:
        base_score = 60.0
    elif emissions < 700:
        base_score = 80.0
    else:
        base_score = 95.0

    final_score = base_score + (coverage * 0.1)
    return round(clamp(final_score, 0, 100), 2)


def calculate_coverage_percent(document_text: str, confidence_percent: float) -> float:
    text_factor = clamp((len(document_text) / 4000.0) * 100, 20, 100)
    coverage = (0.65 * confidence_percent) + (0.35 * text_factor)
    return round(clamp(coverage, 0, 100), 2)


def update_supplier_aggregates(
    supplier_id: str,
    user_email: str,
    incremental_emissions: float,
    current_risk_score: float,
    coverage_percent: float,
    ai_summary: str,
) -> None:
    supplier_ref = db.collection("suppliers").document(supplier_id)
    supplier_snapshot = supplier_ref.get()
    if not supplier_snapshot.exists:
        raise ValueError("Supplier record not found.")

    supplier_data = supplier_snapshot.to_dict() or {}
    if supplier_data.get("user_email") != user_email:
        raise PermissionError("Unauthorized supplier update attempt.")

    existing_documents = list_documents(user_email, supplier_id)
    existing_count = len(existing_documents) - 1
    existing_count = max(existing_count, 0)

    old_total = _coerce_float(supplier_data.get("total_emissions"), 0.0)
    old_coverage = _coerce_float(supplier_data.get("coverage_percent"), 0.0)

    new_count = existing_count + 1
    updated_total = old_total + incremental_emissions

    if new_count <= 1:
        updated_coverage = coverage_percent
    else:
        updated_coverage = ((old_coverage * existing_count) + coverage_percent) / new_count
    updated_risk = calculate_risk_score(updated_total, updated_coverage)

    supplier_ref.update(
        {
            "total_emissions": round(updated_total, 6),
            "risk_score": round(updated_risk, 2),
            "coverage_percent": round(updated_coverage, 2),
            "ai_insight": ai_summary,
        }
    )


def build_dashboard_context(user_email: str) -> Dict[str, Any]:
    suppliers = sorted(
        list_user_suppliers(user_email),
        key=lambda item: _coerce_float(item.get("total_emissions"), 0.0),
    )
    documents = [item for item in list_documents(user_email) if item.get("scope_type", "scope3") == "scope3"]
    verification_queue = _build_verification_queue(user_email)

    total_emissions = sum(float(item.get("total_emissions", 0) or 0) for item in suppliers)
    supplier_count = len(suppliers)
    high_risk_count = len([item for item in suppliers if float(item.get("risk_score", 0) or 0) >= HIGH_RISK_THRESHOLD])
    avg_coverage = (
        round(sum(float(item.get("coverage_percent", 0) or 0) for item in suppliers) / supplier_count, 2)
        if supplier_count
        else 0
    )

    insights = []
    for doc in documents[:3]:
        insights.append(
            {
                "supplier_id": doc.get("supplier_id", ""),
                "supplier_name": doc.get("supplier_name", "Supplier"),
                "emission_total": float(doc.get("extracted_emissions", 0) or 0),
                "coverage_pct": float(doc.get("confidence_percent", 0) or 0),
                "risk_score": calculate_risk_score(
                    float(doc.get("extracted_emissions", 0) or 0),
                    float(doc.get("coverage_percent", doc.get("confidence_percent", 0)) or 0),
                ),
                "risk_level": risk_level(
                    calculate_risk_score(
                        float(doc.get("extracted_emissions", 0) or 0),
                        float(doc.get("coverage_percent", doc.get("confidence_percent", 0)) or 0),
                    )
                ),
                "normalization_alert": "Units standardized to tCO2e.",
                "ai_insight": doc.get("ai_summary", "AI summary unavailable."),
            }
        )

    return {
        "suppliers": suppliers,
        "metrics": {
            "total_emissions": round(total_emissions, 2),
            "supplier_count": supplier_count,
            "high_risk": high_risk_count,
            "avg_coverage": avg_coverage,
            "total_documents": len(documents),
        },
        "insights": insights,
        "verification_queue": verification_queue,
    }

def generate_scope3_report(
    company_name: str,
    supplier: Dict[str, Any],
    documents: List[Dict[str, Any]],
) -> io.BytesIO:
    buffer = io.BytesIO()
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=LETTER,
        leftMargin=32,
        rightMargin=32,
        topMargin=28,
        bottomMargin=28,
    )
    styles = getSampleStyleSheet()

    brand = colors.HexColor("#1A936F")
    brand_dark = colors.HexColor("#0f6f56")
    slate_900 = colors.HexColor("#0f172a")
    slate_700 = colors.HexColor("#334155")
    slate_500 = colors.HexColor("#64748b")
    slate_300 = colors.HexColor("#cbd5e1")
    slate_100 = colors.HexColor("#f1f5f9")
    white = colors.white
    soft_green = colors.HexColor("#e6f5ef")
    soft_amber = colors.HexColor("#fff7ed")
    soft_red = colors.HexColor("#fef2f2")

    title_light = ParagraphStyle(
        "TitleLight",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=26,
        textColor=white,
        spaceAfter=3,
    )
    subtitle_light = ParagraphStyle(
        "SubtitleLight",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#d1fae5"),
    )
    section_style = ParagraphStyle(
        "SectionStyleV2",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=16,
        textColor=slate_900,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "BodyStyleV2",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=13,
        textColor=slate_700,
    )
    meta_style = ParagraphStyle(
        "MetaStyleV2",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=slate_500,
    )
    metric_label_style = ParagraphStyle(
        "MetricLabelV2",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=slate_500,
    )
    metric_value_style = ParagraphStyle(
        "MetricValueV2",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        textColor=slate_900,
    )

    total_emissions = float(supplier.get("total_emissions", 0) or 0)
    risk_score = float(supplier.get("risk_score", 0) or 0)
    coverage = float(supplier.get("coverage_percent", 0) or 0)
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    supplier_name = str(supplier.get("supplier_name", supplier.get("name", "N/A")) or "N/A")
    supplier_code = str(supplier.get("supplier_code", "N/A") or "N/A")
    sector = str(supplier.get("sector", supplier.get("category", "General")) or "General")
    region = str(supplier.get("region", "Global") or "Global")
    risk_label = risk_level(risk_score)

    def _safe_text(value: Any) -> str:
        text = str(value or "")
        return (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
        )

    def _clip(value: Any, limit: int = 95) -> str:
        text = str(value or "").strip().replace("\n", " ")
        if len(text) <= limit:
            return text
        return f"{text[: limit - 3].rstrip()}..."

    story = []

    hero = Table(
        [
            [
                Paragraph("CarbonLens AI Scope 3 Intelligence Report", title_light),
                Paragraph(
                    (
                        f"<b>Generated:</b> {generated_at}<br/>"
                        f"<b>Organization:</b> {_safe_text(company_name or 'N/A')}"
                    ),
                    subtitle_light,
                ),
            ]
        ],
        colWidths=[pdf.width * 0.62, pdf.width * 0.38],
    )
    hero.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), brand),
                ("BOX", (0, 0), (-1, -1), 0, brand),
                ("LEFTPADDING", (0, 0), (-1, -1), 16),
                ("RIGHTPADDING", (0, 0), (-1, -1), 16),
                ("TOPPADDING", (0, 0), (-1, -1), 14),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ]
        )
    )
    story.append(hero)
    story.append(Spacer(1, 12))

    supplier_profile = Table(
        [
            [
                Paragraph("<b>Supplier</b>", meta_style),
                Paragraph(_safe_text(supplier_name), body_style),
                Paragraph("<b>Supplier Code</b>", meta_style),
                Paragraph(_safe_text(supplier_code), body_style),
            ],
            [
                Paragraph("<b>Sector</b>", meta_style),
                Paragraph(_safe_text(sector), body_style),
                Paragraph("<b>Region</b>", meta_style),
                Paragraph(_safe_text(region), body_style),
            ],
        ],
        colWidths=[pdf.width * 0.14, pdf.width * 0.36, pdf.width * 0.14, pdf.width * 0.36],
    )
    supplier_profile.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), slate_100),
                ("GRID", (0, 0), (-1, -1), 0.5, slate_300),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(supplier_profile)
    story.append(Spacer(1, 12))

    risk_bg = soft_green
    risk_color = brand_dark
    if risk_score >= 80:
        risk_bg = soft_red
        risk_color = colors.HexColor("#b91c1c")
    elif risk_score >= 60:
        risk_bg = soft_amber
        risk_color = colors.HexColor("#b45309")

    kpi_cards = Table(
        [
            [
                Paragraph("TOTAL EMISSIONS", metric_label_style),
                Paragraph("RISK LEVEL", metric_label_style),
                Paragraph("COVERAGE", metric_label_style),
                Paragraph("DOCUMENTS", metric_label_style),
            ],
            [
                Paragraph(f"{total_emissions:,.2f} tCO2e", metric_value_style),
                Paragraph(
                    f'<font color="{risk_color}">{_safe_text(risk_label)} ({risk_score:.2f})</font>',
                    metric_value_style,
                ),
                Paragraph(f"{coverage:.2f}%", metric_value_style),
                Paragraph(f"{len(documents)}", metric_value_style),
            ],
        ],
        colWidths=[pdf.width * 0.25, pdf.width * 0.25, pdf.width * 0.25, pdf.width * 0.25],
    )
    kpi_cards.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), slate_100),
                ("BACKGROUND", (1, 1), (1, 1), risk_bg),
                ("BOX", (0, 0), (-1, -1), 0.5, slate_300),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, slate_300),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(kpi_cards)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Document-Level Emissions Analysis", section_style))
    story.append(
        Paragraph(
            (
                "Breakdown of processed evidence with extracted emissions, confidence, and "
                "standardized units used for Scope 3 scoring."
            ),
            body_style,
        )
    )
    story.append(Spacer(1, 6))

    total_doc_emissions = sum(float(item.get("extracted_emissions", 0) or 0) for item in documents)
    doc_table_data = [
        ["Document", "Extracted Emissions", "Contribution", "Category", "Confidence"]
    ]
    if documents:
        for item in documents:
            doc_emissions = float(item.get("extracted_emissions", 0) or 0)
            contribution = (doc_emissions / total_doc_emissions * 100.0) if total_doc_emissions > 0 else 0.0
            doc_table_data.append(
                [
                    _clip(item.get("original_filename", "N/A"), 42),
                    f"{doc_emissions:,.4f} tCO2e",
                    f"{contribution:.2f}%",
                    _clip(item.get("category", "General"), 16),
                    f"{float(item.get('confidence_percent', 0) or 0):.2f}%",
                ]
            )
    else:
        doc_table_data.append(["No documents available", "-", "-", "-", "-"])

    breakdown_table = Table(
        doc_table_data,
        repeatRows=1,
        colWidths=[pdf.width * 0.36, pdf.width * 0.18, pdf.width * 0.14, pdf.width * 0.16, pdf.width * 0.16],
    )
    breakdown_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), brand_dark),
                ("TEXTCOLOR", (0, 0), (-1, 0), white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.8),
                ("GRID", (0, 0), (-1, -1), 0.35, slate_300),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, colors.HexColor("#f8fafc")]),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(breakdown_table)
    story.append(Spacer(1, 14))

    top_drivers: Dict[str, float] = {}
    for doc in documents:
        breakdown_items = doc.get("emission_breakdown") or []
        if not isinstance(breakdown_items, list):
            continue
        for entry in breakdown_items:
            if not isinstance(entry, dict):
                continue
            driver_name = str(entry.get("activity_type", "unknown_activity")).strip() or "unknown_activity"
            kg_value = _coerce_float(entry.get("emissions_kg_co2e"), 0.0)
            if kg_value > 0:
                top_drivers[driver_name] = top_drivers.get(driver_name, 0.0) + (kg_value / 1000.0)
    sorted_drivers = sorted(top_drivers.items(), key=lambda item: item[1], reverse=True)[:6]

    story.append(Paragraph("Top Emission Drivers", section_style))
    if sorted_drivers:
        driver_total = sum(value for _, value in sorted_drivers)
        driver_table = [["Activity", "Emissions (tCO2e)", "Share"]]
        for name, value in sorted_drivers:
            share = (value / driver_total * 100.0) if driver_total > 0 else 0.0
            driver_table.append([_clip(name, 30), f"{value:,.4f}", f"{share:.2f}%"])

        driver_view = Table(
            driver_table,
            repeatRows=1,
            colWidths=[pdf.width * 0.58, pdf.width * 0.22, pdf.width * 0.20],
        )
        driver_view.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), slate_900),
                    ("TEXTCOLOR", (0, 0), (-1, 0), white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, slate_100]),
                    ("GRID", (0, 0), (-1, -1), 0.35, slate_300),
                    ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(driver_view)
    else:
        story.append(Paragraph("No activity-level emission breakdown was available.", body_style))
    story.append(Spacer(1, 14))

    story.append(Paragraph("Emission Reduction Suggestions", section_style))
    manual_suggestions = normalize_reduction_suggestions(supplier.get("reduction_suggestions"))
    if manual_suggestions:
        suggestions_rows = []
        for idx, suggestion in enumerate(manual_suggestions, start=1):
            suggestions_rows.append(
                [
                    Paragraph(f"<b>{idx}.</b>", meta_style),
                    Paragraph(_safe_text(suggestion), body_style),
                ]
            )
        manual_suggestions_table = Table(suggestions_rows, colWidths=[pdf.width * 0.06, pdf.width * 0.94])
        manual_suggestions_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fffc")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#bae6d3")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1fae5")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(manual_suggestions_table)
    else:
        story.append(Paragraph("No manual reduction suggestions are recorded for this supplier yet.", body_style))
    story.append(Spacer(1, 14))

    story.append(Paragraph("AI Recommended Reduction Actions", section_style))
    ai_reduction_actions: List[str] = []
    for driver_name, _driver_value in sorted_drivers[:3]:
        name = str(driver_name).lower()
        if "diesel" in name or "petrol" in name or "gasoline" in name:
            ai_reduction_actions.append(
                "Prioritize fuel optimization by consolidating routes, reducing idle time, and shifting to EV or low-emission fleets."
            )
        elif "electricity" in name or "power" in name:
            ai_reduction_actions.append(
                "Cut electricity emissions through renewable contracts, efficiency retrofits, and smart load scheduling."
            )
        elif "steel" in name or "raw_material" in name:
            ai_reduction_actions.append(
                "Lower material intensity by increasing recycled inputs and setting supplier-level low-carbon material targets."
            )
        elif "freight" in name or "transport" in name:
            ai_reduction_actions.append(
                "Shift freight mix to rail/sea where possible and improve load factor to reduce ton-km emissions."
            )
        else:
            ai_reduction_actions.append(
                f"Create a targeted reduction plan for '{driver_name}' with quarterly milestones and owner accountability."
            )

    if coverage < 70:
        ai_reduction_actions.append(
            "Increase data coverage by collecting more verifiable utility, fuel, and logistics evidence from this supplier."
        )
    if risk_score >= 80:
        ai_reduction_actions.append(
            "Place this supplier on an immediate mitigation track with monthly emissions reviews and executive oversight."
        )

    unique_ai_actions: List[str] = []
    for action in ai_reduction_actions:
        if action not in unique_ai_actions:
            unique_ai_actions.append(action)
    if not unique_ai_actions:
        unique_ai_actions.append(
            "Maintain current controls and keep a continuous monitoring cadence to prevent emissions rebound."
        )

    ai_action_rows = [[Paragraph("&#8226;", meta_style), Paragraph(_safe_text(item), body_style)] for item in unique_ai_actions]
    ai_actions_table = Table(ai_action_rows, colWidths=[pdf.width * 0.04, pdf.width * 0.96])
    ai_actions_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff7ed")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#fed7aa")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#ffedd5")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(ai_actions_table)
    story.append(Spacer(1, 14))

    story.append(Paragraph("AI Insight Highlights", section_style))
    if documents:
        insights_rows: List[List[Any]] = []
        for idx, item in enumerate(documents[:4], start=1):
            insight_text = _clip(item.get("ai_summary", "AI summary unavailable."), 260)
            insights_rows.append(
                [
                    Paragraph(f"<b>Insight {idx}</b>", meta_style),
                    Paragraph(
                        f"<b>{_safe_text(_clip(item.get('original_filename', 'Document'), 55))}</b><br/>"
                        f"{_safe_text(insight_text)}",
                        body_style,
                    ),
                ]
            )
        insights_table = Table(insights_rows, colWidths=[pdf.width * 0.14, pdf.width * 0.86])
        insights_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fffc")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#bae6d3")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1fae5")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(insights_table)
    else:
        story.append(Paragraph("No processed documents available for AI insight generation.", body_style))

    story.append(Spacer(1, 14))
    compliance = Table(
        [
            [
                Paragraph(
                    (
                        "<b>Compliance Note:</b> Generated using AI-assisted extraction with deterministic "
                        "emission factors. All units are normalized to tCO2e."
                    ),
                    meta_style,
                )
            ],
            [Paragraph(f"Timestamp: {generated_at}", meta_style)],
        ],
        colWidths=[pdf.width],
    )
    compliance.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), slate_100),
                ("BOX", (0, 0), (-1, -1), 0.5, slate_300),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(compliance)

    pdf.build(story)
    buffer.seek(0)
    return buffer


@app.errorhandler(413)
def file_too_large(_error):
    flash("Uploaded file is too large.", "error")
    return redirect(url_for("upload"))


@app.route("/")
def index():
    if "user_email" in session:
        user = get_user_profile(session["user_email"])
        if user and user.get("access_role") == "supplier":
            return redirect(url_for("supplier_portal"))
        return redirect(url_for("home"))
    return redirect(url_for("login"))


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "GET":
        return redirect(url_for("login", view="signup"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        name = request.form.get("name", "").strip()
        secret_access_code = request.form.get("secret_access_code", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        company_name = (
            request.form.get("company_name")
            or request.form.get("company")
            or request.form.get("companyName")
            or ""
        ).strip()

        if not all([email, name, password, confirm_password, company_name, secret_access_code]):
            flash("All signup fields are required.", "error")
            return redirect(url_for("login", view="signup"))

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return redirect(url_for("login", view="signup"))

        access_info = _resolve_signup_access(secret_access_code)
        if not access_info:
            flash("Invalid Secret Access Code.", "error")
            return redirect(url_for("login", view="signup"))

        resolved_supplier_id = ""
        if access_info["access_role"] == "supplier":
            resolved_supplier_id = _generate_supplier_id()

        user_ref = db.collection("users").document(email)
        if user_ref.get().exists:
            flash("User already exists.", "error")
            return redirect(url_for("login", view="signup"))

        session["pending_signup"] = {
            "email": email,
            "password": generate_password_hash(password),
            "name": name,
            "company_name": company_name,
            "role": "Staff" if access_info["access_role"] == "staff" else "Supplier",
            "access_role": access_info["access_role"],
            "supplier_id": resolved_supplier_id,
        }

        otp_code = set_otp_session("signup", email)
        if not send_otp_email(email, otp_code, "signup"):
            flash("Failed to send OTP email. Please try again.", "error")
            return redirect(url_for("login", view="signup"))

        flash("OTP sent to your email. Verify to complete signup.", "success")
        return redirect(url_for("otp"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        action = request.form.get("action", "password_login")
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        if not email or not password:
            flash("Email and password are required.", "error")
            return redirect(url_for("login"))

        user_doc = db.collection("users").document(email).get()
        if not user_doc.exists:
            flash("Invalid email or password.", "error")
            return redirect(url_for("login"))

        user_data = user_doc.to_dict() or {}
        if not check_password_hash(user_data.get("password", ""), password):
            flash("Invalid email or password.", "error")
            return redirect(url_for("login"))

        if action == "otp_signin":
            otp_code = set_otp_session("signin", email)
            if not send_otp_email(email, otp_code, "signin"):
                flash("Failed to send OTP email. Please try again.", "error")
                return redirect(url_for("login"))
            flash("OTP sent. Enter the code to sign in.", "success")
            return redirect(url_for("otp"))

        session["user_email"] = email
        access_role = _derive_access_role(user_data)
        if access_role == "supplier":
            return redirect(url_for("supplier_portal"))
        return redirect(url_for("home"))

    if "user_email" in session:
        user = get_user_profile(session["user_email"])
        if user and user.get("access_role") == "supplier":
            return redirect(url_for("supplier_portal"))
        return redirect(url_for("home"))

    initial_view = str(request.args.get("view", "login") or "login").strip().lower()
    if initial_view not in {"login", "signup"}:
        initial_view = "login"
    return render_template("login.html", initial_view=initial_view)


@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully.", "success")
    return redirect(url_for("login"))


@app.route("/resetpwd", methods=["GET", "POST"])
def reset_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not email or not new_password or not confirm_password:
            flash("All fields are required.", "error")
            return redirect(url_for("reset_password"))

        if new_password != confirm_password:
            flash("Passwords do not match.", "error")
            return redirect(url_for("reset_password"))

        user_doc = db.collection("users").document(email).get()
        if not user_doc.exists:
            flash("Account not found.", "error")
            return redirect(url_for("reset_password"))

        session["reset_email"] = email
        session["pending_reset_password"] = generate_password_hash(new_password)
        otp_code = set_otp_session("reset", email)
        if not send_otp_email(email, otp_code, "reset"):
            flash("Failed to send OTP email. Please try again.", "error")
            return redirect(url_for("reset_password"))

        flash("Verification code sent. Enter OTP to finish reset.", "success")
        return redirect(url_for("otp"))

    return render_template("resetpwd.html")


@app.route("/otp", methods=["GET", "POST"])
def otp():
    otp_context = session.get("otp_context")
    otp_email = session.get("otp_email")

    if not otp_context or not otp_email or not session.get("otp_code"):
        return redirect(url_for("login"))

    if request.method == "POST":
        code = request.form.get("otp_code", "").strip()
        if otp_expired():
            flash("OTP expired. Please request a new code.", "error")
            return redirect(url_for("otp"))

        if code != session.get("otp_code"):
            flash("Invalid verification code.", "error")
            return redirect(url_for("otp"))

        if otp_context == "signup":
            pending_signup = session.get("pending_signup") or {}
            if not pending_signup:
                flash("Signup session expired. Please register again.", "error")
                clear_otp_session()
                return redirect(url_for("signup"))

            db.collection("users").document(pending_signup["email"]).set(
                {
                    "email": pending_signup["email"],
                    "password": pending_signup["password"],
                    "name": pending_signup["name"],
                    "company_name": pending_signup["company_name"],
                    "role": pending_signup.get("role", "ESG Manager"),
                    "access_role": pending_signup.get("access_role", "staff"),
                    "supplier_id": pending_signup.get("supplier_id", ""),
                    "created_at": firestore.SERVER_TIMESTAMP,
                }
            )
            session["user_email"] = pending_signup["email"]
            access_role = str(pending_signup.get("access_role", "staff")).strip().lower()
            if access_role == "supplier":
                supplier_public_id = str(pending_signup.get("supplier_id", "") or "").strip()
                if supplier_public_id:
                    db.collection("supplier_profiles").document(supplier_public_id).set(
                        {
                            "supplier_id": supplier_public_id,
                            "supplier_name": pending_signup.get("company_name", pending_signup.get("name", "Supplier")),
                            "user_email": pending_signup["email"],
                            "created_at": firestore.SERVER_TIMESTAMP,
                        },
                        merge=True,
                    )
            session.pop("pending_signup", None)
            clear_otp_session()
            if access_role == "supplier":
                flash(f"Signup verified. Your Supplier ID is {supplier_public_id}. Share it with companies to link records.", "success")
                return redirect(url_for("supplier_portal"))
            flash("Signup verified successfully.", "success")
            return redirect(url_for("home"))

        if otp_context == "signin":
            user = get_user_profile(otp_email)
            session["user_email"] = otp_email
            clear_otp_session()
            flash("OTP verified. Signed in successfully.", "success")
            if user and user.get("access_role") == "supplier":
                return redirect(url_for("supplier_portal"))
            return redirect(url_for("home"))

        if otp_context == "reset":
            email = session.get("reset_email", "")
            hashed_password = session.get("pending_reset_password", "")
            db.collection("users").document(email).update({"password": hashed_password})
            session.pop("reset_email", None)
            session.pop("pending_reset_password", None)
            clear_otp_session()
            flash("Password reset successful. Please sign in.", "success")
            return redirect(url_for("login"))

    return render_template(
        "otp.html",
        destination=otp_email,
        otp_context=otp_context,
        otp_expired=otp_expired(),
    )


@app.route("/otp/resend", methods=["POST"])
def resend_otp():
    otp_context = session.get("otp_context")
    otp_email = session.get("otp_email")

    if not otp_context or not otp_email:
        flash("No active OTP request found.", "error")
        return redirect(url_for("login"))

    otp_code = set_otp_session(otp_context, otp_email)
    if send_otp_email(otp_email, otp_code, otp_context):
        flash("OTP resent successfully.", "success")
    else:
        flash("Failed to resend OTP. Please try again.", "error")

    return redirect(url_for("otp"))


@app.route("/home")
@login_required
def home():
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user:
        session.clear()
        flash("User profile not found. Please log in again.", "error")
        return redirect(url_for("login"))
    if user.get("access_role") == "supplier":
        return redirect(url_for("supplier_portal"))

    context = build_dashboard_context(user_email)

    return render_template(
        "home.html",
        user=user,
        suppliers=context["suppliers"],
        metrics=context["metrics"],
        insights=context["insights"],
        verification_queue=context.get("verification_queue", []),
    )


@app.route("/dashboard")
@login_required
def dashboard():
    user = get_user_profile(session["user_email"])
    if user and user.get("access_role") == "supplier":
        return redirect(url_for("supplier_portal"))
    return redirect(url_for("home"))


@app.route("/suppliers")
@login_required
def suppliers_hub():
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if user and user.get("access_role") == "supplier":
        return redirect(url_for("supplier_portal"))
    suppliers = list_user_suppliers(user_email)
    if suppliers:
        return redirect(url_for("supplier_detail", supplier_id=suppliers[0]["id"]))
    flash("No suppliers found. Add one to get started.", "error")
    return redirect(url_for("add_supplier"))


@app.route("/supplier-portal")
@login_required
def supplier_portal():
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user:
        session.clear()
        flash("Session expired. Please log in again.", "error")
        return redirect(url_for("login"))
    if user.get("access_role") != "supplier":
        return redirect(url_for("home"))

    supplier_external_id = str(user.get("supplier_id", "") or "").strip()
    if not supplier_external_id:
        flash("No supplier ID is linked to this account.", "error")
        return redirect(url_for("login"))

    company_options = list_supplier_company_options(supplier_external_id)
    supplier_records = [item.get("supplier", {}) for item in company_options]

    if supplier_records:
        supplier = dict(supplier_records[0])
    else:
        supplier = {
            "id": supplier_external_id,
            "supplier_code": supplier_external_id,
            "name": user.get("company_name", user.get("name", "Supplier")),
            "supplier_name": user.get("company_name", user.get("name", "Supplier")),
            "total_emissions": 0.0,
            "coverage_percent": 0.0,
            "risk_score": 0.0,
            "category": "General",
            "region": "Global",
        }

    my_total_emissions = sum(_coerce_float(item.get("total_emissions"), 0.0) for item in supplier_records) if supplier_records else 0.0

    company_rollup: Dict[str, Dict[str, Any]] = {}
    all_documents: List[Dict[str, Any]] = []
    for option in company_options:
        company_name = option.get("company_name", "Unknown Company")
        supplier_data = option.get("supplier", {}) if isinstance(option.get("supplier"), dict) else {}
        bucket = company_rollup.setdefault(
            company_name,
            {
                "company_name": company_name,
                "total_emissions": 0.0,
                "supplier_records": 0,
            },
        )
        bucket["total_emissions"] += _coerce_float(supplier_data.get("total_emissions"), 0.0)
        bucket["supplier_records"] += 1

        docs = list_documents_for_supplier(option.get("supplier_id", ""), owner_email=option.get("owner_email") or None)
        for doc_item in docs:
            doc_item["company_name"] = company_name
            all_documents.append(doc_item)

    all_documents.sort(key=lambda item: item.get("created_at_iso", ""), reverse=True)
    docs_count = len(all_documents)
    my_intensity = round(my_total_emissions / docs_count, 2) if docs_count else 0.0

    customer_companies = sorted(company_rollup.values(), key=lambda item: item.get("total_emissions", 0.0), reverse=True)
    companies_served_count = len(customer_companies)
    primary_company = customer_companies[0]["company_name"] if customer_companies else "N/A"

    status_timeline: List[Dict[str, Any]] = []
    for item in all_documents:
        status_timeline.append(
            {
                "filename": item.get("original_filename", "Document"),
                "created_at": item.get("created_at_iso", "")[:10],
                "company_name": item.get("company_name", ""),
                "status_label": _verification_label(str(item.get("verification_status", "verified"))),
                "status_key": str(item.get("verification_status", "verified")).strip().lower(),
                "rejection_reason": item.get("rejection_reason", ""),
                "summary": item.get("ai_summary", ""),
            }
        )

    return render_template(
        "supplier_portal.html",
        user=user,
        supplier=supplier,
        supplier_external_id=supplier_external_id,
        my_total_emissions=my_total_emissions,
        my_intensity=my_intensity,
        customer_companies=customer_companies,
        companies_served_count=companies_served_count,
        primary_company=primary_company,
        status_timeline=status_timeline,
        total_documents=docs_count,
    )


@app.route("/verification/<document_id>/approve", methods=["POST"])
@login_required
def approve_document(document_id: str):
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user or user.get("access_role") != "staff":
        flash("Unauthorized action.", "error")
        return redirect(url_for("home"))

    doc_ref = db.collection("documents").document(document_id)
    doc_snapshot = doc_ref.get()
    if not doc_snapshot.exists:
        flash("Document not found.", "error")
        return redirect(url_for("home"))

    doc_data = doc_snapshot.to_dict() or {}
    if str(doc_data.get("user_email", "")).strip().lower() != user_email.lower():
        flash("You cannot approve this document.", "error")
        return redirect(url_for("home"))

    status = str(doc_data.get("verification_status", "verified")).strip().lower()
    if status == "verified":
        return redirect(url_for("home"))

    supplier_id = str(doc_data.get("supplier_id", "")).strip()
    supplier = get_supplier_owned(supplier_id, user_email)
    if not supplier:
        flash("Supplier not found or unauthorized.", "error")
        return redirect(url_for("home"))

    extracted_emissions = _coerce_float(doc_data.get("extracted_emissions"), 0.0)
    recalculated_payload: Dict[str, Any] = {}
    activity_data = doc_data.get("activity_data") if isinstance(doc_data.get("activity_data"), dict) else {}
    if activity_data:
        try:
            recalculated = calculate_emissions(activity_data)
            extracted_emissions = _coerce_float(recalculated.get("total_tco2e"), extracted_emissions)
            recalculated_payload = {
                "extracted_emissions": extracted_emissions,
                "calculated_emissions_kg_co2e": recalculated.get("total_kg_co2e", 0.0),
                "activity_data": recalculated.get("activity_data", {}),
                "emission_breakdown": recalculated.get("breakdown", []),
            }
        except Exception:
            recalculated_payload = {}

    coverage_percent = _coerce_float(
        doc_data.get("coverage_percent"),
        _coerce_float(doc_data.get("confidence_percent"), 0.0),
    )
    risk_score_value = calculate_risk_score(extracted_emissions, coverage_percent)

    update_supplier_aggregates(
        supplier_id=supplier_id,
        user_email=user_email,
        incremental_emissions=extracted_emissions,
        current_risk_score=risk_score_value,
        coverage_percent=coverage_percent,
        ai_summary=str(doc_data.get("ai_summary", "AI extraction completed.")),
    )
    owner_email = str(supplier.get("user_email", "") or doc_data.get("user_email", "") or user_email).strip()
    trigger_high_emission_alert_if_needed(supplier, owner_email, extracted_emissions)

    approved_doc_updates = {
        "verification_status": "verified",
        "reviewed_by": user_email,
        "reviewed_at": firestore.SERVER_TIMESTAMP,
        "rejection_reason": "",
    }
    approved_doc_updates.update(recalculated_payload)
    doc_ref.update(approved_doc_updates)
    return redirect(url_for("home"))


@app.route("/verification/<document_id>/reject", methods=["POST"])
@login_required
def reject_document(document_id: str):
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user or user.get("access_role") != "staff":
        flash("Unauthorized action.", "error")
        return redirect(url_for("home"))

    doc_ref = db.collection("documents").document(document_id)
    doc_snapshot = doc_ref.get()
    if not doc_snapshot.exists:
        flash("Document not found.", "error")
        return redirect(url_for("home"))

    doc_data = doc_snapshot.to_dict() or {}
    if str(doc_data.get("user_email", "")).strip().lower() != user_email.lower():
        flash("You cannot reject this document.", "error")
        return redirect(url_for("home"))

    rejection_reason = str(request.form.get("rejection_reason", "")).strip()
    doc_ref.update(
        {
            "verification_status": "rejected",
            "rejection_reason": rejection_reason[:500],
            "reviewed_by": user_email,
            "reviewed_at": firestore.SERVER_TIMESTAMP,
        }
    )
    return redirect(url_for("home"))


@app.route("/chat", methods=["GET", "POST"])
@login_required
def carbonlens_chat():
    if request.method == "GET":
        user = get_user_profile(session["user_email"])
        if not user:
            session.clear()
            flash("Session expired. Please log in again.", "error")
            return redirect(url_for("login"))
        return render_template("chat.html", user=user)

    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user:
        return jsonify({"ok": False, "error": "Session expired."}), 401
    payload = request.get_json(silent=True) or {}
    user_message = str(payload.get("message", "")).strip() if payload else ""

    if not user_message:
        return jsonify({"ok": False, "error": "Message is required."}), 400

    chat_history = payload.get("chat_history", [])
    if not isinstance(chat_history, list):
        chat_history = []

    assistant_response = get_carbonlens_ai_response(
        user_message,
        chat_history,
        user_email,
        access_role=str(user.get("access_role", "staff")),
        supplier_id=str(user.get("supplier_id", "")),
    )

    return jsonify(
        {
            "ok": True,
            "reply": assistant_response,
        }
    )


@app.route("/chat/clear", methods=["POST"])
@login_required
def carbonlens_chat_clear():
    return jsonify({"ok": True})


@app.route("/upload", methods=["GET", "POST"])
@login_required
def upload():
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user:
        session.clear()
        flash("Session expired. Please log in again.", "error")
        return redirect(url_for("login"))

    is_supplier_user = str(user.get("access_role", "staff")).strip().lower() == "supplier"
    supplier_company_options = list_supplier_company_options(user.get("supplier_id", "")) if is_supplier_user else []
    if is_supplier_user and not supplier_company_options:
        flash("No company is linked to your Supplier ID yet. Share your Supplier ID with staff to link records.", "error")
        return redirect(url_for("supplier_portal"))
    suppliers = [] if is_supplier_user else list_user_suppliers(user_email)

    if request.method == "POST":
        scope_type = str(request.form.get("scope_type", "scope3") or "scope3").strip().lower()
        if scope_type not in {"scope1", "scope2", "scope3"}:
            flash("Invalid scope selected.", "error")
            return redirect(url_for("upload"))

        if is_supplier_user:
            scope_type = "scope3"

        supplier_id = ""
        selected_supplier_option: Optional[Dict[str, Any]] = None
        if scope_type == "scope3":
            if is_supplier_user:
                requested_supplier_id = str(request.form.get("supplier_id", "") or "").strip()
                if not requested_supplier_id and supplier_company_options:
                    requested_supplier_id = str(supplier_company_options[0].get("supplier_id", "")).strip()
                selected_supplier_option = next(
                    (
                        item
                        for item in supplier_company_options
                        if str(item.get("supplier_id", "")).strip() == requested_supplier_id
                    ),
                    None,
                )
                if not selected_supplier_option:
                    flash("Please select a valid company mapping before upload.", "error")
                    return redirect(url_for("upload"))
                supplier_id = str(selected_supplier_option.get("supplier_id", "")).strip()
            else:
                supplier_id = request.form.get("supplier_id", "").strip()
        submit_mode = request.form.get("submit_mode", "").strip().lower()
        uploaded_file = request.files.get("document") or request.files.get("file")
        manual_field_keys = [
            "diesel_litre",
            "petrol_litre",
            "electricity_kwh",
            "steel_ton",
            "raw_materials_kg",
            "freight_ton_km",
            "rail_freight_ton_km",
        ]
        manual_activity_data: Dict[str, float] = {}
        if not is_supplier_user:
            for key in manual_field_keys:
                value = _coerce_float(request.form.get(key), 0.0)
                if value > 0:
                    manual_activity_data[key] = value
        has_uploaded_file = bool(uploaded_file and uploaded_file.filename and uploaded_file.filename.strip())

        supplier: Optional[Dict[str, Any]] = None
        if scope_type == "scope3":
            if not supplier_id:
                flash("Supplier selection is required for Scope 3.", "error")
                return redirect(url_for("upload"))

            if is_supplier_user:
                supplier = selected_supplier_option.get("supplier") if selected_supplier_option else None
            else:
                supplier = get_supplier_owned(supplier_id, user_email)
            if not supplier:
                flash("Supplier not found or unauthorized.", "error")
                return redirect(url_for("upload"))

        if is_supplier_user and not has_uploaded_file:
            flash("Supplier portal supports document upload only.", "error")
            return redirect(url_for("upload"))

        if not has_uploaded_file and not manual_activity_data:
            flash("Upload a document (PDF/Excel/CSV) or enter manual activity values.", "error")
            return redirect(url_for("upload"))

        try:
            filename = ""
            document_text = ""
            extraction: Dict[str, Any] = {}
            extracted_activity_data: Dict[str, Any] = {}
            category = (
                supplier.get("sector", "General")
                if supplier
                else ("Scope 1" if scope_type == "scope1" else ("Scope 2" if scope_type == "scope2" else "General"))
            )
            confidence = 88
            input_mode = "manual"

            if has_uploaded_file and uploaded_file:
                filename = secure_filename(uploaded_file.filename)
                if not allowed_file(filename):
                    flash("Supported file types: PDF, XLSX, XLSM, CSV.", "error")
                    return redirect(url_for("upload"))

                file_bytes = uploaded_file.read()
                if not file_bytes:
                    flash("Uploaded file is empty.", "error")
                    return redirect(url_for("upload"))

                extension = filename.rsplit(".", 1)[-1].lower()
                if extension == "pdf":
                    document_text = extract_text_from_pdf(file_bytes)
                    extraction = extract_emissions_with_ai(document_text)
                    extracted_activity_data = extraction.get("raw_activity_data") or extraction.get("activity_data") or {}
                    category = extraction.get("category", category)
                    confidence = int(clamp(_coerce_float(extraction.get("confidence_percent"), 50), 1, 100))
                else:
                    extracted_activity_data = extract_activity_data_from_table_file(file_bytes, filename)
                    extraction = {
                        "summary_text": "Spreadsheet extraction completed using deterministic activity mapping.",
                        "category": category,
                        "confidence_percent": 82 if extracted_activity_data else 45,
                    }
                    confidence = int(clamp(_coerce_float(extraction.get("confidence_percent"), 50), 1, 100))
                input_mode = "file"
            else:
                if is_supplier_user:
                    flash("Supplier uploads require document evidence.", "error")
                    return redirect(url_for("upload"))
                timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                filename = f"Manual_Entry_{timestamp}.txt"
                extraction = {
                    "summary_text": "Manual activity entry processed with deterministic emission factors.",
                    "category": category,
                    "confidence_percent": confidence,
                }

            merged_activity_data: Dict[str, float] = {}
            for source in (extracted_activity_data, manual_activity_data):
                if not isinstance(source, dict):
                    continue
                for key, raw_val in source.items():
                    val = _coerce_float(raw_val, 0.0)
                    if val > 0:
                        merged_activity_data[str(key)] = merged_activity_data.get(str(key), 0.0) + val

            if not merged_activity_data:
                raise ValueError("No valid activity values were found in file/manual input.")

            if has_uploaded_file and manual_activity_data:
                input_mode = "file+manual"

            emissions_result = calculate_emissions(merged_activity_data)
            emission_value = emissions_result["total_tco2e"]
            standardized_unit = "tCO2e"

            if input_mode == "manual":
                manual_coverage = (len(emissions_result["activity_data"]) / max(len(EMISSION_FACTORS), 1)) * 100
                coverage = round(clamp(manual_coverage, 40, 95), 2)
            else:
                coverage_text = document_text if document_text else json.dumps(merged_activity_data)
                coverage = calculate_coverage_percent(coverage_text, confidence)

            risk_score_value = calculate_risk_score(emission_value, coverage)
            ai_summary = extraction.get("summary_text", "AI extraction completed.")
            if input_mode == "file+manual":
                ai_summary = f"{ai_summary} Manual activity values were merged with extracted values."
            ai_flag_reason = (
                f"Auto-flagged by AI: extracted {len(emissions_result['activity_data'])} activity metric(s) "
                f"with estimated total {emission_value:,.2f} tCO2e. Staff verification required before finalization."
            )

            owner_email = str(supplier.get("user_email", user_email)).strip() if (is_supplier_user and supplier) else user_email
            verification_status = "waiting_for_staff_approval" if is_supplier_user else "verified"
            supplier_name = (
                supplier.get("supplier_name", supplier.get("name", "Supplier"))
                if supplier
                else str(user.get("company_name", "") or user.get("name", "Company"))
            )

            document_payload = {
                "user_email": owner_email,
                "uploader_email": user_email,
                "uploaded_by_role": user.get("access_role", "staff"),
                "supplier_id": supplier_id,
                "supplier_name": supplier_name,
                "original_filename": filename,
                "storage_path": "",
                "extracted_emissions": emission_value,
                "calculated_emissions_kg_co2e": emissions_result["total_kg_co2e"],
                "standardized_unit": standardized_unit,
                "scope_type": scope_type,
                "category": category,
                "confidence_percent": confidence,
                "coverage_percent": coverage,
                "ai_summary": ai_summary,
                "activity_data": emissions_result["activity_data"],
                "emission_breakdown": emissions_result["breakdown"],
                "emission_factor_dataset": os.path.basename(EMISSION_FACTORS_PATH),
                "file_retained": False,
                "input_mode": input_mode,
                "verification_status": verification_status,
                "rejection_reason": "",
                "ai_flag_reason": ai_flag_reason,
                "created_at": firestore.SERVER_TIMESTAMP,
            }

            document_ref = db.collection("documents").document()
            document_ref.set(document_payload)

            if not is_supplier_user and scope_type == "scope3" and supplier:
                update_supplier_aggregates(
                    supplier_id=supplier_id,
                    user_email=user_email,
                    incremental_emissions=emission_value,
                    current_risk_score=risk_score_value,
                    coverage_percent=coverage,
                    ai_summary=ai_summary,
                )
                trigger_high_emission_alert_if_needed(supplier, owner_email, emission_value)

            db.collection("processing_logs").document().set(
                {
                    "user_email": owner_email,
                    "supplier_id": supplier_id,
                    "document_id": document_ref.id,
                    "scope_type": scope_type,
                    "status": "waiting_for_staff_approval" if is_supplier_user else "completed",
                    "timestamp": firestore.SERVER_TIMESTAMP,
                }
            )

            if is_supplier_user:
                return redirect(url_for("supplier_portal"))
            if scope_type in {"scope1", "scope2"}:
                return redirect(url_for("home"))
            return redirect(url_for("supplier_detail", supplier_id=supplier_id))
        except Exception as exc:
            error_message = str(exc)
            app.logger.exception("Upload processing failed for supplier %s", supplier_id)
            db.collection("processing_logs").document().set(
                {
                    "user_email": user_email,
                    "supplier_id": supplier_id,
                    "document_id": "",
                    "status": "failed",
                    "error_message": error_message[:500],
                    "timestamp": firestore.SERVER_TIMESTAMP,
                }
            )
            flash(f"Processing failed: {error_message}", "error")
            return redirect(url_for("upload"))

    return render_template(
        "upload.html",
        user=user,
        suppliers=suppliers,
        supplier_company_options=supplier_company_options,
        is_supplier_user=is_supplier_user,
        mode="upload",
    )


@app.route("/supplier/<supplier_id>")
@login_required
def supplier_detail(supplier_id: str):
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if user and user.get("access_role") == "supplier":
        return redirect(url_for("supplier_portal"))
    supplier = get_supplier_owned(supplier_id, user_email)

    if not supplier:
        flash("Supplier not found or unauthorized.", "error")
        return redirect(url_for("home"))

    documents = list_documents(user_email, supplier_id)

    breakdown_map: Dict[str, float] = {}
    for item in documents:
        category = item.get("category", "General")
        breakdown_map[category] = breakdown_map.get(category, 0.0) + float(item.get("extracted_emissions", 0) or 0)

    breakdown = [{"label": key, "value": round(val, 4)} for key, val in breakdown_map.items()]
    if not breakdown:
        breakdown = [{"label": supplier.get("sector", "General"), "value": 0.0}]

    supplier["breakdown"] = breakdown
    supplier["documents"] = [
        {
            "name": item.get("original_filename", "N/A"),
            "type": item.get("category", "General"),
            "uploaded_at": item.get("created_at_iso", "")[:10],
            "status": _verification_label(str(item.get("verification_status", "verified"))),
            "proof": item.get("ai_summary", ""),
        }
        for item in documents
    ]

    supplier["ai_insight"] = documents[0].get("ai_summary", supplier.get("ai_insight", "")) if documents else supplier.get("ai_insight", "")
    supplier_risk_score = float(supplier.get("risk_score", 0) or 0)
    supplier_risk_label = risk_level(supplier_risk_score)
    supplier["risk_assessment"] = (
        f"Based on total Scope 3 emissions and coverage of extracted data, this supplier is classified as {supplier_risk_label}."
    )
    supplier["risk_band"] = emissions_risk_band(float(supplier.get("total_emissions", 0) or 0))
    supplier["reduction_suggestions"] = normalize_reduction_suggestions(supplier.get("reduction_suggestions"))

    return render_template(
        "supplier.html",
        user=user,
        supplier=supplier,
        risk_level=supplier_risk_label,
        suppliers=list_user_suppliers(user_email),
    )


@app.route("/supplier/<supplier_id>/suggestions", methods=["POST"])
@login_required
def update_supplier_suggestions(supplier_id: str):
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if user and user.get("access_role") == "supplier":
        return redirect(url_for("supplier_portal"))

    supplier = get_supplier_owned(supplier_id, user_email)
    if not supplier:
        flash("Supplier not found or unauthorized.", "error")
        return redirect(url_for("home"))

    raw_input = request.form.get("reduction_suggestions", "")
    suggestions = normalize_reduction_suggestions(raw_input)
    db.collection("suppliers").document(supplier_id).update({"reduction_suggestions": suggestions})
    flash("Reduction suggestions updated.", "success")
    return redirect(url_for("supplier_detail", supplier_id=supplier_id))


@app.route("/supplier/<supplier_id>/delete", methods=["POST"])
@login_required
def delete_supplier(supplier_id: str):
    user_email = session["user_email"]
    supplier = get_supplier_owned(supplier_id, user_email)
    if not supplier:
        flash("Supplier not found or unauthorized.", "error")
        return redirect(url_for("home"))

    try:
        docs_query = db.collection("documents").where("user_email", "==", user_email).where("supplier_id", "==", supplier_id)
        for doc in docs_query.stream():
            doc.reference.delete()

        logs_query = db.collection("processing_logs").where("user_email", "==", user_email).where("supplier_id", "==", supplier_id)
        for log in logs_query.stream():
            log.reference.delete()

        db.collection("suppliers").document(supplier_id).delete()
        flash("Supplier and related records deleted successfully.", "success")
        return redirect(url_for("home"))
    except Exception:
        flash("Failed to delete supplier.", "error")
        return redirect(url_for("supplier_detail", supplier_id=supplier_id))


@app.route("/add_supplier", methods=["GET", "POST"])
@app.route("/add-supplier", methods=["GET", "POST"])
@login_required
def add_supplier():
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user:
        session.clear()
        flash("Session expired. Please log in again.", "error")
        return redirect(url_for("login"))
    if user.get("access_role") == "supplier":
        return redirect(url_for("supplier_portal"))

    if request.method == "POST":
        supplier_name = (
            request.form.get("supplier_name")
            or request.form.get("name")
            or request.form.get("supplier")
            or ""
        ).strip()
        supplier_code = (
            request.form.get("supplier_code")
            or request.form.get("supplier_id")
            or request.form.get("code")
            or ""
        ).strip()
        sector = (request.form.get("sector") or request.form.get("category") or "General").strip()
        region = (request.form.get("region") or "Global").strip()

        if not supplier_name or not supplier_code:
            flash("Supplier name and supplier code are required.", "error")
            return redirect(url_for("add_supplier"))

        db.collection("suppliers").document().set(
            {
                "user_email": user_email,
                "supplier_name": supplier_name,
                "supplier_code": supplier_code,
                "sector": sector,
                "region": region,
                "risk_score": 0,
                "total_emissions": 0,
                "coverage_percent": 0,
                "reduction_suggestions": [],
                "created_at": firestore.SERVER_TIMESTAMP,
            }
        )
        flash("Supplier added successfully.", "success")
        return redirect(url_for("home"))

    return render_template("addsupp.html", user=user)


@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user:
        session.clear()
        flash("Session expired. Please log in again.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        company_name = (
            request.form.get("company_name")
            or request.form.get("company")
            or user.get("company_name", "")
        ).strip()
        name = (request.form.get("name") or user.get("name", "")).strip()

        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        role = (request.form.get("role") or user.get("role", "ESG Manager")).strip()

        updates: Dict[str, Any] = {
            "company_name": company_name,
            "name": name,
            "role": role,
        }

        if any([current_password, new_password, confirm_password]):
            if not check_password_hash(user.get("password", ""), current_password):
                flash("Current password is incorrect.", "error")
                return redirect(url_for("settings"))

            if not new_password or new_password != confirm_password:
                flash("New passwords do not match.", "error")
                return redirect(url_for("settings"))

            updates["password"] = generate_password_hash(new_password)

        db.collection("users").document(user_email).update(updates)
        flash("Settings updated successfully.", "success")
        return redirect(url_for("settings"))

    return render_template("settings.html", user=user)


@app.route("/settings/delete-account", methods=["POST"])
@login_required
def delete_account():
    user_email = session["user_email"]
    user_doc = db.collection("users").document(user_email).get()
    if not user_doc.exists:
        session.clear()
        flash("Account not found. Please sign in again.", "error")
        return redirect(url_for("login"))

    try:
        documents_query = db.collection("documents").where("user_email", "==", user_email).stream()
        for doc in documents_query:
            doc.reference.delete()

        processing_logs_query = db.collection("processing_logs").where("user_email", "==", user_email).stream()
        for log in processing_logs_query:
            log.reference.delete()

        suppliers_query = db.collection("suppliers").where("user_email", "==", user_email).stream()
        for supplier in suppliers_query:
            supplier.reference.delete()

        db.collection("users").document(user_email).delete()
        session.clear()
        return redirect(url_for("login"))
    except Exception:
        flash("Failed to delete account. Please try again.", "error")
        return redirect(url_for("settings"))


@app.route("/reports")
@login_required
def reports():
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user:
        session.clear()
        flash("Session expired. Please log in again.", "error")
        return redirect(url_for("login"))
    if user.get("access_role") == "supplier":
        return redirect(url_for("supplier_portal"))
    suppliers = list_user_suppliers(user_email)
    return render_template("reports.html", suppliers=suppliers, user=user)


@app.route("/download_report/<supplier_id>")
@login_required
def download_report(supplier_id: str):
    user_email = session["user_email"]
    user = get_user_profile(user_email)
    if not user:
        session.clear()
        flash("Session expired. Please log in again.", "error")
        return redirect(url_for("login"))

    if user.get("access_role") == "supplier":
        linked_options = list_supplier_company_options(user.get("supplier_id", ""))
        linked_supplier = next(
            (item.get("supplier") for item in linked_options if str(item.get("supplier_id", "")).strip() == str(supplier_id).strip()),
            None,
        )
        if not linked_supplier:
            flash("Supplier not found or unauthorized.", "error")
            return redirect(url_for("supplier_portal"))
        supplier = linked_supplier
        owner_email = str(supplier.get("user_email", "") or "").strip()
        documents = list_documents_for_supplier(supplier["id"], owner_email=owner_email)
    else:
        supplier = get_supplier_owned(supplier_id, user_email)
        if not supplier:
            flash("Supplier not found or unauthorized.", "error")
            return redirect(url_for("home"))
        documents = list_documents(user_email, supplier_id)

    if not supplier:
        flash("Supplier not found or unauthorized.", "error")
        return redirect(url_for("home"))

    try:
        pdf_buffer = generate_scope3_report(
            company_name=user.get("company_name", ""),
            supplier=supplier,
            documents=documents,
        )
        filename = f"scope3_report_{secure_filename(supplier.get('supplier_code', supplier_id))}.pdf"
        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )
    except Exception:
        flash("Failed to generate report.", "error")
        return redirect(url_for("supplier_detail", supplier_id=supplier_id))


if __name__ == "__main__":
    app.run(debug=True)

